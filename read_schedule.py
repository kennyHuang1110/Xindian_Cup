import openpyxl
import json

# 讀取 Excel
wb = openpyxl.load_workbook('app/data/新店盃_8隊賽程.xlsx')
print('工作簿:', wb.sheetnames)

# 讀取每個工作表
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n=== {sheet_name} ===')
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True), 1):
        if any(cell is not None for cell in row):
            print(f'{row_idx}: {row}')
